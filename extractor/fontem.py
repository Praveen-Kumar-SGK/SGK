from xlsxwriter.utility import xl_rowcol_to_cell
import tempfile
from .excel_processing import *
import openpyxl
from math import ceil
from openpyxl.utils import range_boundaries
import numpy as np
from decimal import Decimal


# document_location = r"/Users/sakthivel/Documents/SGK/Fontom/"
classifier=joblib.load(fontem_model_location)


def get_input(input_file, input_excel_location):
  if input_file.startswith('\\'):
    print('connecting to SMB share')
    try:
      with smbclient.open_file(r"{}".format(input_file), mode='rb', username=smb_username,
                               password=smb_password) as f:
        with open(input_excel_location, 'wb') as pdf:
          pdf.write(f.read())
        print('file found')
    except:
      smbclient.reset_connection_cache()
      with smbclient.open_file(r"{}".format(input_file), mode='rb', username=smb_username,
                               password=smb_password) as f:
        with open(input_excel_location, 'wb') as pdf:
          pdf.write(f.read())
        print('file found')
    finally:
      smbclient.reset_connection_cache()
    return input_excel_location
  else:
    return document_location + input_file


#filetered row and column wise data
def row_equal_func(columns, rows, df):
    r_slice, c_slice = 0, 0
    rows = list(df.index.values)
    columns = list(df.columns.values)
#     display("1-display",df.head(20))- check
    for col in columns:
        if r_slice == 0:
            for row in rows:
                if str(df[col][row]) in ("blu product") or str(df[col][row]) in ("blu product\n(FRONT)"):
                    c_slice, r_slice = col, row
#                     print("1st_step", df.loc[row])- check
#                     print("b_break",r_slice, c_slice)- check
                    break
#     print("A_break",r_slice, c_slice) - check
    df_slice = df.iloc[r_slice:, c_slice:]
#     start_index = list(df_slice[c_slice].str.contains(r"^blu product").index) - check
    repeat_start_index = df_slice[c_slice].str.contains(r"^blu product")
    repeat_indexes = list(np.flatnonzero(repeat_start_index))
#     print(np.flatnonzero(repeat_start_index))- check
    has_sequence_numbers = False
    for index, value in enumerate(repeat_indexes):
        if index != len(repeat_indexes)-1:
            if value+1 in repeat_indexes:
                continue
            else:
                has_sequence_numbers = False
                break
        else:
            has_sequence_numbers = True
    if has_sequence_numbers:
        df_slice = df_slice.drop_duplicates(subset = c_slice, keep = "last")
#         display("2-display",df_slice.head(20)) - check
    #     df_index_reset = df_slice.reset_index(drop=True)
    df_slice1 = df_slice.dropna(how='all')
#     display("3-display",df_slice1.head(20)) - check
    return df_slice1.T


def data_classification(column_return_df):
    overall_dict = {}
    unwanted = {}
    row_indexes = list(column_return_df.index.values)
    col_indexes = list(column_return_df.columns.values)
    # for _ , value in column_return_df.iterrows():
#     print(column_return_df[col_indexes])
    for _col in col_indexes[:1]:
        for _row in row_indexes:
            key = str(column_return_df[_col][_row])
            rem_squ = re.sub("[\(\[].*?[\)\]]", "", key)
            rem_key = re.sub(r"\d","",rem_squ)
            key_fontem = rem_key.replace("CLP Article","").replace("EUTPD Article","").replace("BACK","").replace("FRONT","").strip()
            key_fontem = key_fontem.replace("SIDE PANEL","").strip()
            x_font = classifier.predict_proba(laser.embed_sentences(key_fontem, lang = "en"))[0]
            x_font.sort()
            classified = classifier.predict(laser.embed_sentences(key_fontem, lang = "en"))[0]
            if x_font[-1]>0.60: # Changed from 70 to 60
                classified_key = classified
            else:
                classified_key = "UNMAPPED"
    #         print(classified_key)
            for inner_col in col_indexes[1:]:
                value = str(column_return_df[inner_col][_row])
                if "intense starter kit" in value.lower().strip() or "blu product" in value.lower().strip() or "market localisation" in value.lower().strip() or "tobacco flavours" in value.lower().strip():
                    unwanted.setdefault("UNMAPPED", []).append({classify(value)[0]:value.strip()})
                else:
                    if value not in ("nan","None","0",):
                        if classified_key in overall_dict:
                            overall_dict[classified_key].append({xl_rowcol_to_cell(inner_col,_row) + "_" +classify(value)[0]:value})
                        else:
                            overall_dict[classified_key] = [{xl_rowcol_to_cell(inner_col,_row) + "_" +classify(value)[0]:value}]
    return overall_dict

def c_round(number, decimal):
    number_of_decimal_places = lambda number: abs(Decimal(str(number)).as_tuple().exponent)
    custom_round = lambda number, decimal: ceil(number * 10 ** decimal) / 10 ** decimal

    if number_of_decimal_places(number) == decimal + 1 and str(number)[-1] == "5":
        return custom_round(number, decimal)
    else:
        return np.round(number, decimal)

def unmergecells_and_assignvalue(ws):
    '''unmerge merged cells and fill all the cells with same value andreturn the same sheet'''
    for merge_cell_range in list(ws.merged_cells):
        min_col, min_row, max_col, max_row= range_boundaries(str(merge_cell_range))
        top_left_cell_value = ws.cell(row=min_row, column=min_col).value            # store merged cell value in string
        ws.unmerge_cells(str(merge_cell_range))                           # unmerge the merged cells
        for row in ws.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):     #iter over merged cells
            for cell in row:               # for each unmerged cells store the merged cell value
                cell.value=top_left_cell_value
    return ws

def convert_to_dataframe(input_excel,sheet_name):
    position = {}
    Excel = openpyxl.load_workbook(input_excel, data_only=True)
    if sheet_name in Excel.sheetnames:
        sheet = Excel[sheet_name]
        sheet = unmergecells_and_assignvalue(sheet)
        final_list = []
        for row in sheet:
            row_list = []
            if sheet.row_dimensions[row[0].row].hidden == False:
                for cell in row:
                    if re.search(r"(\d\.?\d{0,2})\"?(.*)\"?", str(cell.number_format)):
                        decimal_text = re.search(r"(\d\.?\d{0,2})\"?(.*)\"?", str(cell.number_format)).group(1)
                        suffix_text = re.search(r"(\d\.?\d{0,2})\"?(.*)\"?", str(cell.number_format)).group(2)
                        suffix, number_of_decimal_points = None, None
                        if decimal_text:
                            number_of_decimal_points = len(decimal_text.split(".")[-1]) if "." in decimal_text else 0
                            # print(f"{cell.value}-------{cell.number_format}------{number_of_decimal_points}")
                        if suffix_text:
                            suffix = re.sub(r"[^A-Za-z%]", "", suffix_text)
                        if suffix and "%" in suffix:
                            try:
                                if re.search(r"[a-zA-z]", str(cell.value)):
                                    value = cell.value
                                else:
                                    value = f"{cell.value * 100}%"
                            except:
                                value = cell.value
                        else:
    
                            try:
                               
                                if number_of_decimal_points > 0:
                                    value = float(c_round(cell.value, number_of_decimal_points))       # custom round function
                                else:
                                    value = int(c_round(cell.value,0))
                                if suffix:
                                    value = f"{value} {suffix}"
                            except:
                                value = cell.value
                        value = str(value).replace(">","&gt;").replace("<","&lt;")
                        row_list.append(value)
                        position[value] = cell.coordinate
                    else:
                        # print(cell.coordinate, cell.value, f"---{cell.number_format}------")
                        value = str(cell.value).replace(">", "&gt;").replace("<", "&lt;")
                        row_list.append(value)
                        position[value] = cell.coordinate
            final_list.append(row_list)
            # print('-----' * 10)
        df = pd.DataFrame(final_list)
        return df , position


def fontem_main(path,sheetnames):
    temp_directory = tempfile.TemporaryDirectory(dir=document_location)
    input_excel_location = f'{temp_directory.name}/input_excel.xlsx'
    path = get_input(path, input_excel_location)
    final_dict = {}
    for sheet_name in sheetnames.split(","):
        print("filemneame *******",path)
        print('sheetname ********', sheet_name)
        df,_ = convert_to_dataframe(path,sheet_name)
        rows, columns = df.shape
        column_return_df = row_equal_func(columns, rows, df)
        diction = data_classification(column_return_df)
        final_dict[sheet_name] = diction
    return final_dict
