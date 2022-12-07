# viki and sandhya project
import pandas as pd
import openpyxl
from xlsxwriter.utility import xl_col_to_name
from xlsxwriter.utility import xl_rowcol_to_cell
from laserembeddings import Laser
from langid import classify
import time
from sklearn.neural_network import MLPClassifier
import joblib
import io
import re
import smbclient

from environment import MODE

if MODE == 'local':
    from .local_constants import *
else:
    from .dev_constants import *

laser = Laser(path_to_bpe_codes, path_to_bpe_vocab, path_to_encoder)

output_file = io.BytesIO()

def get_file(file):
    if file.startswith('\\'):
        print('connecting to SMB share')
        try:
            with smbclient.open_file(r"{}".format(file), mode='rb', username=smb_username,
                                     password=smb_password) as f:
                output_file.write(f.read())
                output_file.seek(0)
                print('file found')
        except:
            smbclient.reset_connection_cache()
            with smbclient.open_file(r"{}".format(file), mode='rb', username=smb_username,
                                     password=smb_password) as f:
                output_file.write(f.read())
                output_file.seek(0)
                print('file found')
        finally:
            smbclient.reset_connection_cache()
        return 'SMB'
    else:
        return 'LOCAL'

def main(file_name,sheetname):
    out = get_file(file_name)
    if out == 'SMB':
        wb = openpyxl.load_workbook(output_file)
    else:
        file = document_location + file_name
        wb = openpyxl.load_workbook(file)
    try:
        output_file.truncate(0)
    except:
        pass
    try:
        for sheet_in_wb in wb.sheetnames:
            if sheetname in sheet_in_wb:
                sheetname = sheet_in_wb
        sheet = wb[sheetname]
    except:
        return {'status':'0','comment':'sheet name not found'}
    file_name_dic = {}
    # file_name = os.path.basename(path)
    extract_lst = ["FR", "GB", "ES", "IT", "NL", "PL", "RO", "PT", "TR"]
    df = pd.DataFrame(sheet.values)
    rows, columns = df.shape

    def column_equal_func(columns, rows, df):
        r_slice, c_slice = 0, 0
        for col in range(columns):
            for row in range(rows):
                if df[col][row] in extract_lst:
                    c_slice, r_slice = col, row
                    break
        df = df.iloc[r_slice:, c_slice + 1:]
        df = df.dropna(how='all')
        return df

    def row_equal_func(columns, rows, df):
        r_slice,c_slice = 0,0
        for col in range(columns):
            for row in range(rows):
                if df[col][row] in extract_lst:
                    c_slice, r_slice = col, row
                    break
            else:
                continue
            break
        df = df.iloc[r_slice + 1:, c_slice:]
        df = df.dropna(how='all')
        return df

    def colnum_string(n):
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string

    def equal_check_lst(List):
        result = all(element == List[0] for element in List)
        if (result):
            return 1  # all values are equal
        else:
            return 0

    def fun_call_check(columns, rows, df):
        c_slice_lst = []
        # r_slice_lst = []
        for col in range(columns):
            for row in range(rows):
                if df[col][row] in extract_lst:
                    c_slice, r_slice = col, row
                    c_slice_lst.append(c_slice)
        #           r_slice_lst.append(r_slice)

        col_check_lst = equal_check_lst(c_slice_lst)  # def fun
        #############row_check_lst=equal_check_lst(r_slice_lst)
        if col_check_lst == 1:
            column_return_df = column_equal_func(columns, rows, df)  # def fun
            return column_return_df
        else:
            row_return_df = row_equal_func(columns, rows, df)  # def fun
            return row_return_df

    final_df_return = fun_call_check(columns, rows, df)  # def function
    joblib_file = carrefour_model_location
    mlp_model_loaded = joblib.load(joblib_file)
    full_dict = {}
    inner_lst = []
    for index, row in final_df_return.iterrows():
        index = index + 1
        for col_index, col_value in row.items():
            col_alph = colnum_string(col_index + 1)  # def function
            remove_none_val = str(col_value).replace('None', 'n/a').replace('NA', 'n/a').replace('\xa0', ' ').replace(
                '<', '&lt;').replace('>', '&gt;')

            if remove_none_val != 'n/a' and remove_none_val.strip():
                regex_power = "^\d{2,4}\s?[wW]$"
                regex_gram = "^\d{2,4}\s?[gG]$"

                if re.search(regex_power, remove_none_val):
                    if 'TECHNICAL INFORMATION' in full_dict:
                        full_dict['TECHNICAL INFORMATION'].append(
                            {col_alph + str(index) + "_" + language_detection(remove_none_val): remove_none_val})
                    else:
                        full_dict['TECHNICAL INFORMATION'] = [
                            {col_alph + str(index) + "_" + language_detection(remove_none_val): remove_none_val}]

                elif re.search(regex_gram, remove_none_val):
                    if 'TECHNICAL INFORMATION' in full_dict:
                        full_dict['TECHNICAL INFORMATION'].append(
                            {col_alph + str(index) + "_" + language_detection(remove_none_val): remove_none_val})
                    else:
                        full_dict['TECHNICAL INFORMATION'] = [
                            {col_alph + str(index) + "_" + language_detection(remove_none_val): remove_none_val}]


                else:
                    Xtest_laser = laser.embed_sentences(remove_none_val, lang='en')
                    item_res = mlp_model_loaded.predict(Xtest_laser)
                    item_result = item_res[0]

                    # probability
                    item_prob = mlp_model_loaded.predict_proba(Xtest_laser)
                    item_prob[0].sort()
                    prob = item_prob[0][-1]

                    if prob >= 0.85:
                        if item_result in ("DISPLAY MATERIAL", "POWER", "CAPACITY"):
                            item_result = "TECHNICAL INFORMATION"
                        if item_result in full_dict:
                            full_dict[item_result].append(
                                {col_alph + str(index) + "_" + language_detection(remove_none_val): remove_none_val})
                        else:
                            full_dict[item_result] = [
                                {col_alph + str(index) + "_" + language_detection(remove_none_val): remove_none_val}]

                    else:
                        item_result = 'UNMAPPED'
                        if item_result in full_dict:
                            full_dict[item_result].append(
                                {col_alph + str(index) + "_" + language_detection(remove_none_val): remove_none_val})
                        else:
                            full_dict[item_result] = [
                                {col_alph + str(index) + "_" + language_detection(remove_none_val): remove_none_val}]

    file_name_dic[file_name] = full_dict
    # print(file_name_dic)
    return full_dict

from .excel_processing import *
import langid
import langdetect as lang_det

def language_detection(text):
    print("text-------->",text)
    if isinstance(text,str) and str(text).strip():
        text = str(text).replace("\n"," ").strip()
    else:
        return "en"
    language_set = ['en', 'fr', 'es','nl','it','pl','ro','pt']
    langid.set_languages(['en', 'fr', 'es','nl','it','pl','ro','pt'])
    # with GoogleTranslate(text) as out:
    #     if out["language"] in language_set:
    #         return out["language"]
    try:
        fasttext_output = language_model.predict_pro(text)[0]
        # print(f'fasttext---->{fasttext_output}')
        if fasttext_output[0] in language_set:
            if fasttext_output[1] > 0.50:
                return fasttext_output[0]
        langid_output = classify(text)[0]
        # print(f'langid---->{langid_output}')
        if langid_output in language_set:
            if langid_output == fasttext_output[0]:
                return langid_output
        langdetect_output = lang_det.detect_langs(text)[0]
        # print(f'langdetect---->{langdetect_output}')
        langdetect_lang , lang_detect_prob = str(langdetect_output).split(':')
        if langdetect_lang in language_set:
            if float(lang_detect_prob) > 0.70:
                return langdetect_lang
        return classify(text)[0]
    except:
        return "en"

def carrefour_main(filepath,sheetnames):    # for sheet loop single url
    final_dict = {}
    for sheet in sheetnames.split(","):
        sheet_response = main(filepath,sheet)
        final_dict[sheet] = sheet_response
    return final_dict

# final_response = test(sheet)
# print(final_response)
# Stop = time.time()
# print("\nTime Taken to execute : ", Stop - Start)
