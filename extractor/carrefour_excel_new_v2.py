import openpyxl
import io

from .excel_processing import *

import warnings
warnings.filterwarnings("ignore")



# path_to_bpe_codes = r'/opt/anaconda3/lib/python3.8/site-packages/laserembeddings/data/93langs.fcodes'
# path_to_bpe_vocab = r'/opt/anaconda3/lib/python3.8/site-packages/laserembeddings/data/93langs.fvocab'
# path_to_encoder = r'/opt/anaconda3/lib/python3.8/site-packages/laserembeddings/data/bilstm.93langs.2018-12-26.pt'

# laser = Laser(path_to_bpe_codes, path_to_bpe_vocab, path_to_encoder)
# document_location = r"/Users/sakthivel/Documents/SGK/Carrefour/"

from environment import MODE
if MODE == 'local':
    from .local_constants import *
else:
    from .dev_constants import *

output_file = io.BytesIO()

def get_file(file):
    if file.startswith('\\'):
        print('connecting to SMB share')
        try:
            with smbclient.open_file(r"{}".format(file), mode='rb', username=smb_username,
                                     password=smb_password) as f:
                output_file.write(f.read())
                output_file.seek(0)
                print('file found')
        except:
            smbclient.reset_connection_cache()
            with smbclient.open_file(r"{}".format(file), mode='rb', username=smb_username,
                                     password=smb_password) as f:
                output_file.write(f.read())
                output_file.seek(0)
                print('file found')
        finally:
            smbclient.reset_connection_cache()
        return 'SMB'
    else:
        return 'LOCAL'

def main(file_name, sheetname):
    out = get_file(file_name)
    if out == 'SMB':
        wb = openpyxl.load_workbook(output_file)
    else:
        file = document_location + file_name
        wb = openpyxl.load_workbook(file)
    try:
        output_file.truncate(0)
    except:
        pass
    try:
        for sheet_in_wb in wb.sheetnames:
            if sheetname in sheet_in_wb:
                sheetname = sheet_in_wb
        sheet = wb[sheetname]
    except:
        return {'status': '0', 'comment': 'sheet name not found'}
    file_name_dic = {}
    # file_name = os.path.basename(path)
    extract_lst = ["FR", "GB", "ES", "IT", "NL", "PL", "RO", "PT", "TR"]
    df = pd.DataFrame(sheet.values)
    rows, columns = df.shape

    def column_equal_func(columns, rows, df):
        r_slice, c_slice = 0, 0
        for col in range(columns):
            for row in range(rows):
                if df[col][row] in extract_lst:
                    c_slice, r_slice = col, row
                    break
        df = df.iloc[r_slice:, c_slice + 1:]
        df = df.dropna(how='all')
        return df.T  # Included Transpose Method

    def row_equal_func(columns, rows, df):
        r_slice, c_slice = 0, 0
        for col in range(columns):
            for row in range(rows):
                if df[col][row] in extract_lst:
                    c_slice, r_slice = col, row
                    break
            else:
                continue
            break
        df = df.iloc[r_slice + 1:, c_slice:]
        df = df.dropna(how='all')
        return df

    def colnum_string(n):
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string

    def equal_check_lst(List):
        result = all(element == List[0] for element in List)
        if (result):
            return 1  # all values are equal
        else:
            return 0

    def fun_call_check(columns, rows, df):
        c_slice_lst = []
        # r_slice_lst = []
        for col in range(columns):
            for row in range(rows):
                if df[col][row] in extract_lst:
                    c_slice, r_slice = col, row
                    c_slice_lst.append(c_slice)
        #           r_slice_lst.append(r_slice)

        col_check_lst = equal_check_lst(c_slice_lst)  # def fun
        #############row_check_lst=equal_check_lst(r_slice_lst)
        if col_check_lst == 1:
            print("************* Column wise format")
            column_return_df = column_equal_func(columns, rows, df)  # def fun
            #             return column_return_df
            overall_list = []
            for index, row in column_return_df.iterrows():
                index = index + 1
                #             print(index,"*******",row)
                row_list = []
                for col_index, col_value in row.items():
                    #                 print(col_index,"&&&&&&&", col_value)
                    col_alph = colnum_string(index)  # def function
                    remove_none_val = str(col_value).replace('None', 'n/a').replace('NA', 'n/a').replace('\xa0',
                                                                                                         ' ').replace(
                        '<', '&lt;').replace('>', '&gt;')
                    if remove_none_val.strip() not in ['n/a', '']:
                        row_list.append({col_alph + str(col_index + 1) + "_" + classify(remove_none_val)[
                            0]: remove_none_val.strip()})
                if row_list:
                    overall_list.append(row_list)

            return overall_list
        else:
            print("************* Row wise format")

            row_return_df = row_equal_func(columns, rows, df)  # def fun
            #             return row_return_df
            overall_list = []
            for index, row in row_return_df.iterrows():
                index = index + 1
                #             print(index,"*******",row)
                row_list = []
                for col_index, col_value in row.items():
                    #                 print(col_index,"&&&&&&&", col_value)
                    col_alph = colnum_string(col_index + 1)  # def function
                    remove_none_val = str(col_value).replace('None', 'n/a').replace('NA', 'n/a').replace('\xa0',
                                                                                                         ' ').replace(
                        '<', '&lt;').replace('>', '&gt;')
                    if remove_none_val.strip() not in ['n/a', '']:
                        row_list.append(
                            {col_alph + str(index) + "_" + classify(remove_none_val)[0]: remove_none_val.strip()})
                if row_list:
                    overall_list.append(row_list)
            return overall_list

    final_list_of_list = fun_call_check(columns, rows, df)  # def function
    print(final_list_of_list)
    return final_list_of_list


def get_single_dimension_array(list):
    content_list  = [content for inner_dict in list for _ , content in inner_dict.items()]
    multi_dimensional_array = laser.embed_sentences(content_list,lang="en")
    single_dimension = multi_dimensional_array.mean(axis=0).reshape(1,1024)
    return single_dimension

model = joblib.load(carrefour_model_location)
def predict(input,threshold=0.70,ignore_element:tuple=("None")):
    predicted_class = model.predict(input)[0]
    predicted_class_probability = model.predict_proba(input)
    predicted_class_probability[0].sort()
    max_predicted_class_probability = max(predicted_class_probability[0])
    print("class_predicted---->", input, "--------->", predicted_class,'------->',max_predicted_class_probability)
    if max_predicted_class_probability > threshold and predicted_class not in ignore_element:
        return predicted_class
    else:
        return "UNMAPPED"

def carrefour_extraction(file_name, sheetname):
    sheet_dict = {}
    list_of_list = main(file_name,sheetname)
    for row_list in list_of_list:
        if len(row_list) < 3:
            predicted_class = "UNMAPPED"
        else:
            single_dimension_array = get_single_dimension_array(row_list)
            predicted_class = predict(input=single_dimension_array)
        if predicted_class == "UNMAPPED" and len(row_list) > 3:
            predicted_class = "PRODUCT_FEATURES"
        for inner_dict in row_list:
            for cell_value , content in inner_dict.items():
                if predicted_class in sheet_dict:
                    sheet_dict[predicted_class].append({cell_value:content})
                else:
                    sheet_dict[predicted_class] =  [{cell_value:content}]
    return sheet_dict

def carrefour_main(file_name, sheetnames):
    final_dict = {}
    for sheet_name in sheetnames.split(","):
        response = carrefour_extraction(file_name,sheet_name)
        final_dict[sheet_name] = response
    return final_dict



# out = main("BBM550-21_checkQLK_multi.xlsx","REFERENCE")
