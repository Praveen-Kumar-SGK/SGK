import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from langid import classify
import string
import re
import joblib
from .excel_processing import *
import tempfile

# --------------------------------------------------------------------------------
# Laser Embedding
from laserembeddings import Laser
# path_to_bpe_codes = r'/Users/manirathinams/opt/anaconda3/lib/python3.9/site-packages/laserembeddings/data/93langs.fcodes'
# path_to_bpe_vocab = r'/Users/manirathinams/opt/anaconda3/lib/python3.9/site-packages/laserembeddings/data/93langs.fvocab'
# path_to_encoder = r'/Users/manirathinams/opt/anaconda3/lib/python3.9/site-packages/laserembeddings/data/bilstm.93langs.2018-12-26.pt'
laser = Laser(path_to_bpe_codes, path_to_bpe_vocab, path_to_encoder) 

#classifier model
classifier=joblib.load(purina_excel_model_loc)

#document_location='/Users/manirathinams/Documents/KT/Excel/Purina/Purina/excel.xlsx'

# --------------------------------------------------------------------------------

def get_input(input_file, input_excel_location):
  if input_file.startswith('\\'):
    print('connecting to SMB share')
    try:
      with smbclient.open_file(r"{}".format(input_file), mode='rb', username=smb_username,
                               password=smb_password) as f:
        with open(input_excel_location, 'wb') as pdf:
          pdf.write(f.read())
        print('file found')
    except:
      smbclient.reset_connection_cache()
      with smbclient.open_file(r"{}".format(input_file), mode='rb', username=smb_username,
                               password=smb_password) as f:
        with open(input_excel_location, 'wb') as pdf:
          pdf.write(f.read())
        print('file found')
    finally:
      smbclient.reset_connection_cache()
    return input_excel_location
  else:
    return document_location + input_file

# --------------------------------------------------------------------------------

def extracting_worksheet(file, sheet_name):
    doc_format=file.split('.')[-1].lower()
    if doc_format in ('xlsx', 'xlsm', 'xls'):     # Checking file format
        wb=load_workbook(file, data_only=True)
        sheet=wb.sheetnames   
        if sheet_name in sheet:         # Checking sheetname with available sheetnames
            ws=wb[sheet_name]
        else:
            raise Exception("Invalid sheet name")
    else:
        raise Exception("Invalid file format is not supported")
    return ws

# --------------------------------------------------------------------------------
#function for bold extraction

def bold_conversion(ws):
    for i in range(1, ws.max_row+1):
        for j in range(1, ws.max_column+1):
            if (ws.cell(i,j).font.bold==True) and (ws.cell(i,j).value!=None):
                ws[ws.cell(i,j).coordinate] = '&lt;b&gt;' + ws.cell(i,j).value +'&lt;/b&gt;' 

# --------------------------------------------------------------------------------
#function for percentage 

def percentage_extraction(ws):
    for i in range(1, ws.max_row+1):
        for j in range(1, ws.max_column+1):
            if ws.cell(i,j).number_format in ['0%','0.0%','0.00%']:
                ws[ws.cell(i,j).coordinate]= str(ws.cell(i,j).value*100)+"%"

# --------------------------------------------------------------------------------
# extracting the contents with coordinates

def extracting_values_and_coordinates(file, sheet_name): 
    ws=extracting_worksheet(file, sheet_name)
#    ws=unmerge_rows(ws)   #if need to run unmergecells_and_assignvalue function uncomment this line
    bold_conversion(ws)
    percentage_extraction(ws)
    
    Newlist=[]
    for i in range(1, ws.max_row+1):
        temp=[]
        for j in range(1, ws.max_column+1):
            #if ws.cell(i,j).value!=None:
            #below line append both cell coordinates and values in a list, if don't need remove a "ws.cell[i,j].coordinate"
            temp.append([ws.cell(i,j).coordinate, str(ws.cell(i,j).value).strip()])
        if temp:
            Newlist.append(temp)
    df=pd.DataFrame(Newlist)
    return df

# --------------------------------------------------------------------------------
def removing_unwanted_cols(df):
    
    unwanted_cols=['Uploaded file', 'Position', 'Field type', 'Attachment', 'Comments', 'Category', 'Content type', 
          'Source language', 'Comments (for translation)', 'Comments (text source only)']

    #code to get a coordinates like AA, AB, BC
    required_col=[i[0][:-1] for i in df.iloc[0,:] if i]
    df=df.iloc[:,:len(required_col)]
    df.columns=required_col

    #appending the unwanted column coordinates in unwnt_col_indx
    unwnt_col_indx=[]
    for x in range(0, len(df.columns)):
        if df.iloc[0,x]:
            if df.iloc[0,x][1].replace('&lt;b&gt;','').replace('&lt;/b&gt;','').strip() in unwanted_cols: 
                unwnt_col_indx.append(df.columns[x])

    #removing the unwanted columns from dataframe using unwnt_col_indx list
    dframe=[]
    for i in range(len(df)):
        row=[]
        for j in range(len(df.columns)):
            if df.iloc[i,j]:
                if re.sub('[0-9]','',df.iloc[i,j][0]) not in unwnt_col_indx:  #checking column coordinate like(AA,BD) with unwnt_col_indx 
                    row.append(df.iloc[i,j])
        if row:
            dframe.append(row)

    data=pd.DataFrame(dframe)
    return data

# --------------------------------------------------------------------------------
def header_cleaning(data):
    #removing the coordinates in 1st column which is header
    data[data.columns[0]] = data[data.columns[0]].apply(lambda cell : cell[1])

    #forward fill the values in 1st colummn
    for i in range(len(data)):
        if data.iloc[i,0]=='None':
            data.iloc[i,0]=data.iloc[i-1,0]

    #removing the cells with None values from dataframe
    frame=[]
    for s in range(len(data)):
        row=[]
        for t in range(len(data.columns)):
            if data.iloc[s,t][1]!='None':
                row.append(data.iloc[s,t])
        if row:
            frame.append(row)

    dfs=pd.DataFrame(frame)
    dfs=dfs.fillna('')
    return dfs

# --------------------------------------------------------------------------------
#Extracing the json format with key header classification
def key_val(dfs):
    key=[]
    val=[]
    for s in range(len(dfs)):
        temp=(dfs.iloc[s,0].split('-'))[0]
        new_string= re.sub(r'[0-9]+', '', temp) 
        ky_cln=new_string.replace('&lt;b&gt;','').replace('&lt;/b&gt;','').replace('(','').replace(')','') .strip()       
        proba=classifier.predict_proba(laser.embed_sentences(ky_cln, lang='en'))[0]
        proba.sort()
        content=[]
        for t in range(1, len(dfs.columns)): 
            if dfs.iloc[s,t]!='':
                content.append({str(dfs.iloc[s,t][0])+'_'+str(classify(str(dfs.iloc[s,t][1]))[0]):dfs.iloc[s,t][1]})
        if content:
            val.append(content)
            if proba[-1] >0.70:
                key.append((classifier.predict(laser.embed_sentences(ky_cln, lang='en')))[0])
            else:
                key.append('UNMAPPED')
    #Creating dictionary using key, value pair
    general={}
    for x in range(len(key)):
        general.setdefault(key[x],[]).extend(val[x])   
    return general

# --------------------------------------------------------------------------------
#main function

def purina_main(file, sheetnames):
    temp_directory = tempfile.TemporaryDirectory(dir=document_location)
    input_excel_location = f'{temp_directory.name}/input_excel.xlsx'
    file = get_input(file, input_excel_location)
    
    final_dict = {}
    for sheet_name in sheetnames.split(","):
        df=extracting_values_and_coordinates(file, sheet_name)
        data=removing_unwanted_cols(df)
        dfs=header_cleaning(data)
        general=key_val(dfs)
        final_dict[sheet_name] = general
        return final_dict
# --------------------------------------------------------------------------------
