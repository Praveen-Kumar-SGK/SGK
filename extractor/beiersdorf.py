# Packages used
from openpyxl import load_workbook
from .excel_processing import *
import tempfile
import joblib 

# document_location = r"/Users/sakthivel/Documents/SGK/Beiersdorf/Beiersdorf/Samples/"
classifier =joblib.load(beiersdorf_model_location)
#--------------------------------------------------------------------------------
def get_input(input_file, input_excel_location):
  if input_file.startswith('\\'):
    print('connecting to SMB share')
    try:
      with smbclient.open_file(r"{}".format(input_file), mode='rb', username=smb_username,
                               password=smb_password) as f:
        with open(input_excel_location, 'wb') as pdf:
          pdf.write(f.read())
        print('file found')
    except:
      smbclient.reset_connection_cache()
      with smbclient.open_file(r"{}".format(input_file), mode='rb', username=smb_username,
                               password=smb_password) as f:
        with open(input_excel_location, 'wb') as pdf:
          pdf.write(f.read())
        print('file found')
    finally:
      smbclient.reset_connection_cache()
    return input_excel_location
  else:
    return document_location + input_file

#--------------------------------------------------------------------------------
# Extracting Data
def data_extraction(path,sheet):
    wb=load_workbook(path,data_only=True)
    ws=wb[sheet]
    all_rows=[]
    for row in ws.iter_rows(min_row=1):
        row_temp=[]
        for cell in row:
            if cell.value != None and str(cell.value).strip()!='':
                #print(cell.value)
                # if (cell.value).startswith('='):
                #     # print(cell.coordinate)
                #     wb1=load_workbook(path,data_only=True)
                #     ws1=wb1[sheet]
                #     # print(ws1[cell.coordinate].value)
                #     a=([(str(cell.coordinate)+'_'+str(classify(str(cell.value))[0])),str(ws1[cell.coordinate].value)])
                if cell.font.bold:
                    a=[(str(cell.coordinate)+'_'+str(classify(str(cell.value))[0])),"&lt;b&gt;"+str(cell.value)+"&lt;/b&gt;"]
                else:
                    a=[(str(cell.coordinate)+'_'+str(classify(str(cell.value))[0])),str(cell.value).replace('<','&lt;').replace('>','&gt;')]
                row_temp.append(a)
        all_rows.append(row_temp)
    return all_rows

#--------------------------------------------------------------------------------
# Selecting the particular columns and rows from the column names
def sliced_rows(all_rows):
    index=[]
    main_rows=[]
    for i in range(len(all_rows)):
        # print('\n',i[0])
        for j in all_rows[i]:
            key=re.sub(' +', ' ',j[1]).lower().replace("&lt;/b&gt;",'').replace("&lt;b&gt;",'')
            #if 'current aw' in key or 'proposed aw' in key or 'current india maw english' in key or 'proposed india aw english' in key or 'current india aw english' in key:
            if "aw" in key and ('current' in key or 'proposed' in key):
                index.append(j[0][0]) #j[0][0] will contain the column name for which the value is needed
                main_rows=all_rows[i+1:] # slicing the rows from the index of required columns
    return main_rows,index
#--------------------------------------------------------------------------------
def key_value_rows(main_rows,index):
    char_list=list(string.ascii_uppercase) #list of alphabets
    # taking the key columns name before the value columns is found
    key_col=[char_list[i-1] for i in range(len(char_list)) if char_list[i]==index[0]]
    temp=[]
    row_cleaned=[]
    # Some rows might not have keys (e.g Front Panel), we need to map it the previous key until new key comes
    for row in main_rows:
        r=[]
        for cell in row:
            if cell[0][0] in key_col:
                temp=cell
                # print(temp)
            elif cell[0][0] in index:
                r.append([temp,cell])
                # print(r)
        if r:
            row_cleaned.append(r)
    return row_cleaned
#--------------------------------------------------------------------------------
def value_extraction(row_cleaned):
    values=[]
    keys=[]
    for row in row_cleaned:
        # print(row)
        for cells in row:
            for i in range(len(cells)):
                if i==1:
                    temp, prob, classified = cleaning_classifiction(cells[i][1], classifier)
                    values.append([{cells[i][0]:cells[i][1]}])
                    # print(cells[i][1])
                    if prob[-1]>0.7:
                        keys.append(classified)
                    else:
                        keys.append("MARKETING_CLAIM")
    
    return keys,values
#--------------------------------------------------------------------------------
def cleaning_classifiction(key_id, classifier):
    temp = key_id.replace("&lt;b&gt;", '').replace("&lt;/b&gt;", '').replace("&lt;/b&gt;", '').replace('&lt;','').replace('&gt;', '').strip()
    # temp=re.sub('[\W_]+',' ',temp)
    prob = classifier.predict_proba(laser.embed_sentences(temp.strip(), lang='en'))[0]
    prob.sort()
    classified = str(classifier.predict(laser.embed_sentences(temp, lang='en'))[0])
    return temp, prob, classified
#--------------------------------------------------------------------------------
def multi_key_value(keys,values):
    file = {}
    for i in range(len(keys)):
        file.setdefault(keys[i], []).extend(values[i])
    return file
#--------------------------------------------------------------------------------

def beiersdorf_main(path,sheetnames):
    temp_directory = tempfile.TemporaryDirectory(dir=document_location)
    input_excel_location = f'{temp_directory.name}/input_excel.xlsx'
    path = get_input(path,input_excel_location)
    final_dict = {}
    for sheet_name in sheetnames.split(","):
      print('sheetname', sheet_name)
      all_rows=data_extraction(path,sheet_name)
      main_rows,index=sliced_rows(all_rows)
      if main_rows:
          row_cleaned=key_value_rows(main_rows, index)
          keys,values=value_extraction(row_cleaned)
          final_dict[sheet_name]=multi_key_value(keys, values)
      else:
          final_dict[sheet_name] = {}
    return final_dict



