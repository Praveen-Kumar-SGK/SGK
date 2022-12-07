# ---------------------------------------- Goodman Fielder -----------------------------------------#
# Packages used
from openpyxl import load_workbook
import tempfile
from .utils import GetInput
from .excel_processing import *

# # --------------------------------------------------------------------------------
# # Laser Embedding
# path_to_bpe_codes = r'/opt/anaconda3/lib/python3.8/site-packages/laserembeddings/data/93langs.fcodes'
# path_to_bpe_vocab = r'/opt/anaconda3/lib/python3.8/site-packages/laserembeddings/data/93langs.fvocab'
# path_to_encoder = r'/opt/anaconda3/lib/python3.8/site-packages/laserembeddings/data/bilstm.93langs.2018-12-26.pt'
# laser = Laser(path_to_bpe_codes, path_to_bpe_vocab, path_to_encoder)

# Loading data
# Path1 = "/Users/praveen/Documents/Study/Projects/Goodman/Samples/"
# --------------------------------------------------------------------------------
# Loading MLP Classifier for classification keys
classifier = joblib.load(goodman_fielder_model)
# --------------------------------------------------------------------------------
# Nutrition Table Keys
nutrition_keys = ['Energy', 'Protein', 'Total Fat', 'Saturated Fat', 'Carbohydrate', 'Sugar', 'Total Fibre',
                  'Soluble Fibre',
                  'Fructans', 'Beta Glucan', 'Insoluble Fibre', 'Resistant Starch', 'Sodium', 'Potassium', 'Fibre',
                  'Vitamin B1', 'Vitamin B3',
                  'Vitamin B6', 'Vitamin E', 'Iron', 'Zinc']
# --------------------------------------------------------------------------------
# Extracting Data
def data_extraction(sheet_name, filename):
    location = filename
    wb = load_workbook(location)
    print("sheet_name---->",sheet_name)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        main_rows = []
        for row in ws.iter_rows(min_row=1):
            # print(row)
            row_temp = []
            for cell in row:
                if cell.value != None and str(cell.value).strip() != '':
                    # print(cell.value)
                    if cell.font.bold:
                        a = {(str(cell.coordinate) + '_' + str(classify(str(cell.value))[0])): "&lt;b&gt;" + str(
                            cell.value) + "&lt;/b&gt;"}
                    else:
                        a = {(str(cell.coordinate) + '_' + str(classify(str(cell.value))[0])): str(cell.value).replace('<',
                                                                                                                       '&lt;').replace(
                            '>', '&gt;')}
                    row_temp.append(a)
                # print(row_temp)
            if row_temp:
                main_rows.append(row_temp)
        return main_rows, location
    else:
        raise Exception("sheet name does not exist")


# --------------------------------------------------------------------------------
# Cleaning and Classification
def cleaning_classifiction(key_id):
    temp = list(key_id[0].values())[0].replace("&lt;b&gt;", '').replace("&lt;/b&gt;", '').replace("&lt;/b&gt;",
                                                                                                  '').replace('&lt;',
                                                                                                              '').replace(
        '&gt;', '').strip()
    temp = re.sub('[\W_]+', ' ', temp)
    prob = classifier.predict_proba(laser.embed_sentences(temp.strip(), lang='en'))[0]
    prob.sort()
    classified = str(classifier.predict(laser.embed_sentences(temp, lang='en'))[0])
    return temp, prob, classified


# --------------------------------------------------------------------------------
# Extracting data from rows which has single value
def single_value_rows(main_rows):
    keys = []
    values = []
    mutli_values = []
    for i in range(len(main_rows)):
        if len(main_rows[i]) == 1:
            temp, prob, classified = cleaning_classifiction(main_rows[i])
            # Nutrition Table Content
            if 'Percentage Daily Intakes' in temp:
                keys.append('NUTRITION_TABLE_CONTENT')
                values.append(main_rows[i])
            # Multi columns values
            elif list(main_rows[i][0].keys())[0][0] in ['B', 'C', 'D']:
                if list(main_rows[i][0].keys())[0][0] != 'A' and list(main_rows[i - 1][0].keys())[0][0] == 'A':
                    # Tagging the previous key to the value which is not aligned with the key
                    temp1, prob1, classified1 = cleaning_classifiction(main_rows[i - 1])
                    if prob1[-1] > 0.65:
                        keys.append(classified1)
                        values.append([main_rows[i][0]])
                        mutli_values.append(classified1)
                    else:
                        keys.append('UNMAPPED')
                        values.append([main_rows[i][0]])
                        mutli_values.append('UNMAPPED')
                elif list(main_rows[i][0].keys())[0][0] != 'A':  # Aligning Multiple values to key
                    keys.append(mutli_values[0])
                    values.append(main_rows[i])
            # Ingredients
            elif classified == 'INGREDIENTS_DECLARATION':
                if prob[-1] > 0.9:  # Higher probability cause of misclassification
                    keys.append(classified)
                    values.append(main_rows[i])
                else:
                    keys.append('UNMAPPED')
                    values.append(main_rows[i])
            elif list(main_rows[i][0].keys())[0][0] == 'A':
                mutli_values = []
                keys.append('UNMAPPED')
                values.append([main_rows[i][0]])
            else:
                keys.append('UNMAPPED')
                values.append([main_rows[i][0]])

    return keys, values


# --------------------------------------------------------------------------------
# Extracting data from rows which has key and value
def key_value_rows(main_rows):
    keys = []
    values = []
    for i in range(len(main_rows)):
        if len(main_rows[i]) == 2:
            temp, prob, classified = cleaning_classifiction(main_rows[i])
            if prob[-1] > 0.65:
                keys.append(classified)
                values.append([main_rows[i][1]])
            else:
                keys.append('UNMAPPED')
                values.append([main_rows[i][1]])
    return keys, values


# --------------------------------------------------------------------------------
# Extracting data from rows which has key and multiple values
def multivalues_rows(main_rows):
    keys = []
    values = []
    nkeys = []
    nvalues = []
    for i in range(len(main_rows)):
        if len(main_rows[i]) > 2:
            nut_temp = []
            for j in main_rows[i]:
                nut_temp.append(j)
            temp, prob, classified = cleaning_classifiction(nut_temp)
            # Checking for Nutrition Key and values
            try:
                # Checking if its a nutrition key and 2nd element is a value
                second_element = list(nut_temp[1].values())[0].replace("&lt;b&gt;", '').replace("&lt;/b&gt;",
                                                                                                '').replace(
                    "&lt;/b&gt;", '').replace('&lt;', '').replace('&gt;', '').strip()
                if prob[-1] > 0.65 and classified in nutrition_keys and float(second_element) or int(
                        second_element) == 0:
                    nkeys.append(classified)
                    ntemp = []
                    for j in nut_temp[1:]:
                        if list(j.values())[0] in ['kJ', 'mg', 'g', 'KJ', '%', '%RDI**']:  # Finding the units
                            ntemp.append({'Unit': j})
                        else:
                            ntemp.append({'Value': j})  # Finding the values
                    if len(ntemp) == (len(nut_temp) - 1):
                        nvalues.append(ntemp)

            except ValueError:
                # Serving Size
                if prob[-1] > 0.65 and classified == 'SERVING_SIZE':
                    if len(nut_temp) > 2:
                        for k in range(len(nut_temp[1:])):  # Capturing values
                            keys.append(classified)
                            values.append([main_rows[i][k + 1]])
                    else:
                        keys.append(classified)
                        values.append([main_rows[i][1]])
                else:
                    for j in nut_temp:
                        keys.append('UNMAPPED')
                        values.append([j])

    return keys, values, nkeys, nvalues


# --------------------------------------------------------------------------------
# Final Output
def multi_key_value(k, v):
    f = {}
    for i in range(len(k)):
        if k[i] not in f:
            f[k[i]] = [v[i][0]]
        else:
            if type(v[i]) != list:
                f[k[i]] = [v[i][0]]
            else:
                f[k[i]].append(v[i][0])
    return f


# --------------------------------------------------------------------------------
# Main function
def main(filename, sheet_names):
    temp_directory = tempfile.TemporaryDirectory(dir=document_location)
    doc_format = os.path.splitext(filename)[1].lower()
    input_excel_location = f'{temp_directory.name}/input_excel{doc_format}'
    get_input = GetInput(filename, input_excel_location)
    filename = get_input()
    sheet_dict = {}
    for sheet_name in sheet_names.split(','):
        main_rows, location = data_extraction(sheet_name, filename)
        keys1, values1 = single_value_rows(main_rows)
        keys2, values2 = key_value_rows(main_rows)
        keys3, values3, nkeys, nvalues = multivalues_rows(main_rows)
        # Combining all the keys and values
        final_keys = keys1 + keys2 + keys3
        final_values = values1 + values2 + values3
        final = multi_key_value(final_keys, final_values)

        # Nutrition Facts Dictionary
        NUTRITION_FACTS = {}
        for i in range(len(nkeys)):
            NUTRITION_FACTS[nkeys[i]] = nvalues[i]
        final['NUTRITION_FACTS'] = [NUTRITION_FACTS]
        sheet_dict[sheet_name] = final
    return sheet_dict


# --------------------------------------------------------------------------------
# JSON Format output
# filename = "TAB-177917 Nam Duong Hang Viet Soy Sauce.xlsx"  # str(input('Enter file name:'))
# sheet_name = 'Technical Artwork Brief'  # str(input('Enter sheet name: '))
# output = main(filename, sheet_name)
# --------------------------------------------------------------------------------
